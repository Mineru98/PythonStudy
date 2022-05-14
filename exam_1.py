import openpyxl
from tqdm import tqdm
import argparse
import os
from pprint import pprint

def process(companyName, sheetName, inputFileName, outFileName, startRowIdx, startColIdx):
    wb = openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.realpath(__file__)), "assets", inputFileName)) # 엑셀 파일 불러오기
    sheet = wb[sheetName] # 원하는 시트 선택하기

    result_wb = openpyxl.Workbook()
    result_sheet = result_wb.active

    columns = [] # 칼럼명 리스트

    for row in sheet.iter_rows(min_row=int(startRowIdx), max_row=int(startRowIdx), min_col=int(startColIdx)):
        idx = 1
        for cell in row:
            columns.append(cell.value)
            result_sheet.cell(column=idx, row=1, value=cell.value)
            idx += 1

    rows = []
    for row in tqdm(sheet.iter_rows(min_row=int(startRowIdx)+1, min_col=int(startColIdx))):
        data = {}
        for idx in range(len(row)):
            data[columns[idx]] = row[idx].value
        rows.append(data)

    result = []

    for row in rows:
        if companyName is None:
            if row["입금일자"] is None or row["입금일자"] == "":
                result.append(row)
        else:
            if str(row["상호"]).find(companyName) != -1:
                if row["입금일자"] is None or row["입금일자"] == "":
                    result.append(row)

    for idx in range(len(result)):
        for c_idx in range(len(columns)):
            result_sheet.cell(column=c_idx + 1, row=idx+1+1, value=result[idx][columns[c_idx]])
    result_wb.save(os.path.join(os.path.dirname(os.path.realpath(__file__)), "assets", outFileName))


def showSheetName(inputFileName):
    wb = openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.realpath(__file__)), "assets", inputFileName)) # 엑셀 파일 불러오기
    pprint(wb.sheetnames)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(prog="엑셀 자동화", description="설명")
    parser.add_argument("-i", dest="input_file_name", help="가져올 엑셀 파일명", required=True)
    parser.add_argument("-o", dest="out_file_name", help="내보낼 엑셀 파일명", required=False)
    parser.add_argument("-s", dest="sheet_name", help="시트이름", required=False)
    parser.add_argument("-sr", dest="start_row_idx", help="시작줄", required=False)
    parser.add_argument("-sc", dest="start_col_idx", help="시작 칼럼 위치", required=False)
    parser.add_argument("-c", dest="company", help="원하는 회사명", required=False)
    parser.add_argument("-m", dest="mode", help="모드", required=False)

    args = parser.parse_args()
    input_file_name = args.input_file_name
    out_file_name = args.out_file_name
    sheet_name = args.sheet_name
    start_row_idx = args.start_row_idx
    start_col_idx = args.start_col_idx
    company = args.company
    mode = args.mode

    if mode is None or mode == "1":
        # 돈 안받은 내역 추출
        process(companyName=company, sheetName=sheet_name, inputFileName=input_file_name, outFileName=out_file_name, startRowIdx=start_row_idx, startColIdx=start_col_idx)
    elif mode == "2":
        # 시트 이름 리스트 보기
        showSheetName(inputFileName=input_file_name)
